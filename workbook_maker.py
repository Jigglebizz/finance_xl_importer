import math
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import List
import copy
import pdb
from statement import Statement, Categorizer
from bank import BankInfo, BankType

#------------------------------------------------------------------------------------------------
class ExcelColumn:
  name : str

  def __init__( self, name :str ) -> None:
    self.name = name

  def NameToInt( self, name : str ) -> int:
    row_idx = 1
    i = 0
    for c in name:
      row_idx += ( ord(c) - 65 ) + i * 26
      i += 1
    return row_idx
  
  def IntToName( self, i : int ) -> str:
    name : str = ''
    while i != 0:
      name = chr(((i-1) % 26)+65) + name
      i = math.floor( (i - 1 ) / 26 )
    return name
  
  def __add__( self, num : int ):
    i_rep = self.NameToInt( self.name )
    i_rep += num
    return ExcelColumn( self.IntToName( i_rep ) )
  
  def __sub__( self, num : int ):
    i_rep = self.NameToInt( self.name )
    i_rep -= num
    return ExcelColumn( self.IntToName( i_rep ) )
  
  def __gt__( self, other ) -> bool:
    return self.NameToInt( self.name ) > other.NameToInt( other.name )

  def __repr__( self ) -> str:
    return self.name
  
#------------------------------------------------------------------------------------------------
class ExcelCell:
  col : ExcelColumn
  row : int

  def __init__( self, col : ExcelColumn, row : int ) -> None:
    self.col = col
    self.row = row

  def inc_col( self ):
    ret = ExcelCell( copy.deepcopy( self.col ), self.row )
    self.col += 1
    return ret

  def __repr__( self ) -> str:
    return f'{str(self.col)}{self.row}'
  

#------------------------------------------------------------------------------------------------
class ExcelCursor:
  cell_cursor : ExcelCell
  cell_begin  : ExcelCell
  width       : int

  def __init__( self, cell : ExcelCell, width : int ) -> None:
    self.cell_begin  = copy.deepcopy( cell )
    self.cell_cursor = copy.deepcopy( cell )
    self.width       = width
  
  def inc( self ) -> ExcelCell:
    ret_cell = copy.deepcopy( self.cell_cursor )
  
    cell_end : ExcelCell = ExcelCell( self.cell_begin.col + self.width - 1, self.cell_cursor.row )
    self.cell_cursor.col += 1
    if self.cell_cursor.col > cell_end.col:
      self.cell_cursor.col = copy.deepcopy( self.cell_begin.col )
      self.cell_cursor.row += 1
  
    return ret_cell


#------------------------------------------------------------------------------------------------
class WorkbookMaker:
  path     : str
  workbook : openpyxl.Workbook

  def __init__( self, path ) -> None:
    self.path     = path
    self.workbook = openpyxl.load_workbook( path )

  def Save( self ) -> None:
    self.workbook.save( self.path )

  def AppendStatements( self, new_statements : List[ Statement ], categorizer : Categorizer ) -> None:
    min_start_date = min( [ stmt.start_date for stmt in new_statements ] )
    max_end_date   = max( [ stmt.end_date   for stmt in new_statements ] )
    start_date_str = min_start_date.strftime( '%b%d %Y' )
    end_date_str   = max_end_date.strftime( '%b%d %Y')

    worksheet_name = f'{ start_date_str }-{ end_date_str }'
    if worksheet_name in self.workbook.sheetnames:
      self.workbook.remove( self.workbook[ worksheet_name ] )

    sheet = self.workbook.create_sheet( worksheet_name )
    simpl_name = worksheet_name.replace(' ', '').replace('-','_')

    transaction_table_name = f'{simpl_name}_tx'
    cell_cursor : ExcelCell = self.MakeSummaryTable( name          = f'{simpl_name}_summary',
                                                     start_cell    = ExcelCell( ExcelColumn( 'A'), 1 ),
                                                     sheet         = sheet,
                                                     tx_table_name = transaction_table_name, 
                                                     statements    = new_statements,
                                                     categorizer   = categorizer )
    cell_cursor.col += 2
    cell_cursor.row  = 1

    cell_cursor = self.MakeTransactionTable( name          = transaction_table_name, 
                                             start_cell    = cell_cursor,
                                             sheet         = sheet, 
                                             statements    = new_statements )

  #------------------------------------------------------------------------------------------------
  # returns the bottom right cell
  def MakeTransactionTable( self, name: str, start_cell: ExcelCell, sheet : openpyxl.worksheet.worksheet, statements : List[ Statement ] ) -> ExcelCell:
    all_tx = [ tx for stmt in statements for tx in stmt.transactions  ]
    all_tx.sort( key=lambda x : x.date, reverse=True)

    sheet[ str( start_cell ) ] = 'Transactions'
    start_cell.row += 1
    cursor : ExcelCursor = ExcelCursor( start_cell, 8 )

    table_start = start_cell
    table_end   = None

    sheet[ str( cursor.inc() ) ] = 'Date'
    sheet[ str( cursor.inc() ) ] = 'Amount'
    sheet[ str( cursor.inc() ) ] = 'Account'
    sheet[ str( cursor.inc() ) ] = 'Account Type'
    sheet[ str( cursor.inc() ) ] = 'Description'
    sheet[ str( cursor.inc() ) ] = 'Category'
    sheet[ str( cursor.inc() ) ] = 'Matched Transfer'
    sheet[ str( cursor.inc() ) ] = 'Is Debt'
    for tx in all_tx:
      sheet[ str( cursor.inc() ) ] = tx.date.strftime( '%m/%d/%Y' )

      amt_cell = cursor.inc()
      sheet[ str( amt_cell ) ] = float(tx.amount.AsExcel())
      sheet[ str( amt_cell ) ].number_format = "$0.00"

      sheet[ str( cursor.inc() ) ] = str( tx.account )
      sheet[ str( cursor.inc() ) ] = str( tx.account.bank_info.type.name )
      sheet[ str( cursor.inc() ) ] = tx.name
      sheet[ str( cursor.inc() ) ] = tx.category

      sheet[ str( cursor.inc() ) ] = 'TRUE' if tx.matched_transfer else 'FALSE'
      table_end = cursor.inc()
      sheet[ str( table_end ) ] = 'TRUE' if tx.is_debt else 'FALSE'

    stmt_table  = Table( displayName=name, ref=f'{str( table_start ) }:{ str( table_end ) }' )
    table_style = TableStyleInfo( name='TableStyleMedium9', showRowStripes=True )
    stmt_table.tableStyleInfo = table_style
    sheet.add_table( stmt_table )
  
  #------------------------------------------------------------------------------------------------
  # returns the bottom right cell
  def MakeSummaryTable( self, name : str, start_cell : ExcelCell, sheet : openpyxl.worksheet.worksheet, tx_table_name : str, statements : List[ Statement ], categorizer : Categorizer ) -> ExcelCell:
    sheet[ str( start_cell ) ] = 'Summary'
    cell = copy.deepcopy( start_cell )
    cell.row += 1

    table_start = cell
    table_end   = None
    cursor = ExcelCursor( cell, len( statements ) + 2 )

    sheet[ str( cursor.inc() ) ] = 'Category'
    sheet[ str( cursor.inc() ) ] = 'Totals'
    for stmt in statements:
      sheet[ str( cursor.inc() ) ] = str( stmt.account )

    sheet[ str( cursor.inc() ) ] = 'Debt Change'
    sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Is Debt], "*TRUE*" )'
    for stmt in statements:
      sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Is Debt], "*TRUE*", {tx_table_name}[Account], "*{str(stmt.account)}*")'
    
    sheet[ str( cursor.inc() ) ] = 'Expenses'
    sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Amount], "<0", {tx_table_name}[Matched Transfer], "*FALSE*")'
    for stmt in statements:
      sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Amount], "<0", {tx_table_name}[Account], "*{str(stmt.account)}*", {tx_table_name}[Matched Transfer], "*FALSE*")'

    sheet[ str( cursor.inc() ) ] = 'Total Revenue'
    sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Amount], ">0", {tx_table_name}[Matched Transfer], "*FALSE*" )'
    for stmt in statements:
      sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Amount], ">0", {tx_table_name}[Account], "*{str(stmt.account)}*" )'

    sheet[ str( cursor.inc() ) ] = 'Cash Flow'
    sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Matched Transfer], "*FALSE*", {tx_table_name}[Is Debt], "*FALSE*" )'
    for stmt in statements:
      sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Account], "*{str(stmt.account)}*", {tx_table_name}[Is Debt], "*FALSE*")'

    sheet[ str( cursor.inc() ) ] = 'Net Change'
    sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Matched Transfer], "*FALSE*" )'
    for stmt in statements:
      sheet[ str( cursor.inc() ) ] = f'=SUMIFS({tx_table_name}[Amount], {tx_table_name}[Account], "*{str(stmt.account)}*" )'

    for cat in categorizer.categories:
      sheet[ str( cursor.inc() ) ] = cat
      sheet[ str( cursor.inc() ) ] = f'=SUMIFS( {tx_table_name}[Amount], {tx_table_name}[Category], "*{cat}*" )'
      for stmt in statements:
        table_end = cursor.inc()
        sheet[ str( table_end ) ] = f'=SUMIFS( {tx_table_name}[Amount], {tx_table_name}[Category], "*{cat}*", {tx_table_name}[Account], "*{str(stmt.account)}*" )'

    stmt_table  = Table( displayName=name, ref=f'{str( table_start ) }:{ str( table_end ) }' )
    table_style = TableStyleInfo( name='TableStyleMedium9', showRowStripes=True )
    stmt_table.tableStyleInfo = table_style
    sheet.add_table( stmt_table )

    return table_end